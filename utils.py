"""
Export helpers for Dataloom v3.0.

Owns CSV and Excel (xlsx) generation from query results. Intended for use
by the download endpoint; has no side effects beyond producing bytes.

Public API: export_csv, export_excel, make_export_filename
"""

import csv
import io
from datetime import datetime


def export_csv(headers: list, records: list) -> bytes:
    """Convert query result headers and records into a UTF-8 CSV byte string.

    Accepts records in either dict or list form so callers can pass the
    ``d.records`` payload from ``/api/query`` directly without reshaping.

    Args:
        headers: Ordered list of column names. Defines column order for
            both the header row and dict-based record extraction.
        records: Rows to write. Each row may be either a ``dict`` keyed by
            header name, or a plain ``list`` of values in column order.

    Returns:
        UTF-8 BOM-encoded CSV bytes ready to be sent as a file download.
        The BOM (utf-8-sig) ensures Excel opens the file without a codec
        dialog on Windows.
    """
    if headers is None:
        raise ValueError("export_csv() received None for headers")
    if records is None:
        raise ValueError("export_csv() received None for records")
    headers = list(headers)
    records = list(records)
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)

    writer.writerow(headers)

    for row in records:
        if isinstance(row, dict):
            writer.writerow([row.get(h, "") for h in headers])
        else:
            writer.writerow(row)

    # utf-8-sig adds a BOM that signals UTF-8 to Excel on Windows
    return output.getvalue().encode("utf-8-sig")


def export_excel(headers: list, records: list, sheet_name: str = "Results") -> bytes:
    """Convert query result headers and records into an .xlsx byte string.

    Produces a styled workbook with a formatted header row, alternating row
    shading, auto-sized columns (sampled from the first 50 rows), a frozen
    header row, and a second "Info" sheet with export metadata.

    Raises ``RuntimeError`` with an actionable install message rather than
    an ``ImportError`` if openpyxl is missing, so the error surfaces cleanly
    through the API response.

    Args:
        headers: Ordered list of column names.
        records: Rows to write. Each row may be either a ``dict`` keyed by
            header name, or a plain ``list`` of values in column order.
        sheet_name: Title of the primary data sheet. Truncated to 31
            characters to satisfy Excel's sheet-name limit. Defaults to
            ``"Results"``.

    Returns:
        Raw ``.xlsx`` bytes ready to be sent as a file download.

    Raises:
        RuntimeError: If openpyxl is not installed.
    """
    if headers is None:
        raise ValueError("export_excel() received None for headers")
    if records is None:
        raise ValueError("export_excel() received None for records")
    headers = list(headers)
    records = list(records)
    sheet_name = sheet_name or "Results"
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError(
            "openpyxl is not installed. Run: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None  # openpyxl always sets active sheet on a new Workbook
    # Excel enforces a 31-character sheet name limit
    ws.title = sheet_name[:31]

    # ── Header row styling ────────────────────────────────────────
    # Deep indigo fill and light text matches the Dataloom UI accent palette
    HEADER_FILL  = PatternFill("solid", fgColor="1E1B4B")
    HEADER_FONT  = Font(bold=True, color="C7D2FE", size=11)
    HEADER_ALIGN = Alignment(horizontal="left", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # ── Data rows ─────────────────────────────────────────────────
    ROW_FONT  = Font(size=10)
    ROW_ALIGN = Alignment(horizontal="left", vertical="center")

    for row_idx, row in enumerate(records, start=2):
        if isinstance(row, dict):
            values = [row.get(h, "") for h in headers]
        else:
            values = list(row)

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = ROW_FONT
            cell.alignment = ROW_ALIGN

        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    "solid", fgColor="F8F9FF"
                )

    # ── Auto-size columns ─────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        # Sample max content width from header + first 50 rows to keep
        # this O(columns) rather than O(rows) for large result sets
        max_len = len(str(header))
        for row_idx in range(2, min(len(records) + 2, 52)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"

    # ── Add metadata in a second sheet ───────────────────────────
    meta = wb.create_sheet("Info")
    meta["A1"] = "Generated by"
    meta["B1"] = "Dataloom v3.0"
    meta["A2"] = "Exported at"
    meta["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    meta["A3"] = "Rows"
    meta["B3"] = len(records)
    meta["A4"] = "Columns"
    meta["B4"] = len(headers)
    meta.column_dimensions["A"].width = 16
    meta.column_dimensions["B"].width = 28

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def make_export_filename(format: str = "csv") -> str:
    """Generate a timestamped export filename for a download response.

    Args:
        format: File extension without a leading dot (e.g. ``"csv"`` or
            ``"xlsx"``). Defaults to ``"csv"``.

    Returns:
        Filename string in the form ``dataloom_export_YYYYMMDD_HHMMSS.<format>``.
    """
    format = format or "csv"
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"dataloom_export_{ts}.{format}"
